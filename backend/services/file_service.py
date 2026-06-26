import os
import re
import uuid
import shutil
import logging
from datetime import datetime, timezone
from pypinyin import lazy_pinyin
from flask import current_app, request, send_file
from models import db
from models.user import User, ROLE_ADMIN, PERM_UPLOAD, PERM_DOWNLOAD, PERM_COPY
from models.file import Directory, FileRecord
from models.log import AuditLog, CopyAudit
from utils.validators import allowed_file, get_file_extension, sanitize_filename
from markupsafe import escape
from utils.request import get_client_ip, get_client_ua

logger = logging.getLogger(__name__)


def _audit(user, action, detail, commit=True):
    """Write a file-module audit log entry.

    Args:
        user: User object (provides id, account, username).
        action: Short action identifier (e.g. 'upload', 'delete').
        detail: Human-readable description of the operation.
        commit: If True, commit immediately; set False when the caller
                needs to batch the audit write with other changes.
    """
    log = AuditLog(
        user_id=user.id, account=user.account, username=user.username,
        ip=get_client_ip() if request else 'system', user_agent=get_client_ua() if request else '',
        module='file', action=action, detail=detail,
    )
    db.session.add(log)
    if commit:
        db.session.commit()


def _detach_copy_audits(*file_ids):
    """Clear copy audit FK refs so file records can be permanently deleted."""
    ids = [i for i in file_ids if i is not None]
    if not ids:
        return
    CopyAudit.query.filter(CopyAudit.file_id.in_(ids)).update(
        {CopyAudit.file_id: None}, synchronize_session=False
    )


def _safe_user_dir(user):
    """Generate a safe directory name for a user: pinyin_username + user.id."""
    raw = user.username or str(user.id)
    if re.search(r'[\u4e00-\u9fff]', raw):
        raw = ''.join(lazy_pinyin(raw))
    safe = re.sub(r'[^\w\-.]', '_', raw)
    return f"{safe}_{user.id}"


def _has_public_permission(user, directory_id, perm_flag):
    """Check if user has a public-scope permission.

    Lookup order:
    1. Exact directory permission (directory_id)
    2. Default public permission (directory_id=None)
    Admins bypass the check.
    """
    if user.is_admin():
        return True
    from models.user import UserPermission
    # 1. directory-specific permission
    perm = UserPermission.query.filter_by(
        user_id=user.id, scope='public', directory_id=directory_id
    ).first()
    if perm and perm.has_permission(perm_flag):
        return True
    # 2. default public permission (fallback)
    perm = UserPermission.query.filter_by(
        user_id=user.id, scope='public', directory_id=None
    ).first()
    return bool(perm and perm.has_permission(perm_flag))


def _resolve_path(relative_path):
    """Resolve a relative storage path to absolute path on disk."""
    return os.path.normpath(os.path.join(current_app.config['UPLOAD_FOLDER'], relative_path))


def _physical_path(scope, user, filename):
    """Generate physical storage path for a file (returns relative path).

    Returns relative path (relative to UPLOAD_FOLDER):
        public/  → public/{16-hex}_{filename}
        private/ → private/{pinyin_account}_{user.id}/{16-hex}_{filename}
    """
    safe_name = sanitize_filename(filename)
    unique_name = f"{uuid.uuid4().hex[:16]}_{safe_name}"
    if scope == 'public':
        return f'public/{unique_name}'
    else:
        user_dir_name = _safe_user_dir(user)
        abs_user_dir = os.path.join(current_app.config['PRIVATE_FOLDER'], user_dir_name)
        os.makedirs(abs_user_dir, exist_ok=True)
        return f'private/{user_dir_name}/{unique_name}'


def _ensure_user_root_dir(user):
    """Get or create the user's private root directory in the database.

    Each user has a dedicated root directory under /private/{safe_user_dir}.
    Returns the Directory object.
    """
    user_dir_name = _safe_user_dir(user)
    path = f'/private/{user_dir_name}'
    d = Directory.query.filter_by(path=path).first()
    if d:
        return d
    d = Directory(
        name=user_dir_name,
        parent_id=None,
        scope='private',
        owner_id=user.id,
        path=path,
        created_by=user.id,
    )
    db.session.add(d)
    db.session.flush()
    return d


def init_root_directories():
    """Create root public and private directories if they don't exist."""
    for scope in ('public', 'private'):
        path = f'/{scope}'
        if not Directory.query.filter_by(path=path).first():
            d = Directory(name=scope, scope=scope, path=path)
            db.session.add(d)
    db.session.commit()


def upload_file(user, files, directory_id=None, scope='public'):
    """Upload one or multiple files. Returns (success, result)."""
    if not files:
        return False, '未选择文件'

    # Permission check
    if scope == 'public' and not _has_public_permission(user, directory_id, PERM_UPLOAD):
        return False, '没有上传权限'

    max_size = current_app.config.get('SINGLE_FILE_MAX_SIZE', 50 * 1024 * 1024)
    batch_max = current_app.config.get('BATCH_MAX_SIZE', 200 * 1024 * 1024)
    directory = Directory.query.get(directory_id) if directory_id else None

    # Pre-check total batch size by seeking to end of each file to get byte count.
    total_size = 0
    for f in files:
        f.seek(0, 2)  # seek to end
        total_size += f.tell()  # current position = file size in bytes
        f.seek(0)  # reset to beginning for actual save
    if total_size > batch_max:
        return False, f'批量上传总大小超过限制（最大{batch_max // 1024 // 1024}MB）'

    success_count = 0
    fail_count = 0
    results = []

    # For private uploads without a specific directory, resolve user root once
    user_root_dir = None
    if scope == 'private' and directory_id is None:
        user_root_dir = _ensure_user_root_dir(user)

    for f in files:
        filename = f.filename
        if not filename:
            fail_count += 1
            results.append({'filename': filename, 'status': 'failure', 'reason': '文件名为空'})
            continue

        if not allowed_file(filename):
            fail_count += 1
            results.append({'filename': filename, 'status': 'failure', 'reason': '文件格式不允许'})
            continue

        # Read file to check size
        f.seek(0, 2)
        size = f.tell()
        f.seek(0)

        if size > max_size:
            fail_count += 1
            results.append({'filename': filename, 'status': 'failure',
                            'reason': f'文件大小超过限制（最大{max_size // 1024 // 1024}MB）'})
            continue

        actual_scope = scope
        owner_id = user.id

        rel_path = _physical_path(actual_scope, user, filename)
        abs_path = _resolve_path(rel_path)
        os.makedirs(os.path.dirname(abs_path), exist_ok=True)
        f.save(abs_path)

        effective_directory_id = directory.id if directory else None
        if actual_scope == 'private' and effective_directory_id is None and user_root_dir:
            effective_directory_id = user_root_dir.id

        ext = get_file_extension(filename)
        record = FileRecord(
            filename=os.path.basename(rel_path),
            original_filename=filename,
            file_path=rel_path,
            file_size=size,
            file_type=ext,
            mime_type=f.content_type,
            directory_id=effective_directory_id,
            owner_id=owner_id,
            scope=actual_scope,
        )
        db.session.add(record)
        db.session.flush()
        success_count += 1
        results.append({'filename': filename, 'status': 'success', 'file_id': record.id})

    db.session.commit()

    dir_name = directory.name if directory else '根目录'
    _audit(user, 'upload',
           f'上传了 {success_count} 个文件到「{dir_name}」' + (f'（{fail_count} 个失败）' if fail_count else ''))

    return True, {
        'success_count': success_count,
        'fail_count': fail_count,
        'results': results,
    }


def download_file(user, file_id):
    """Download a file. Returns (success, result_or_error)."""
    record = FileRecord.query.get(file_id)
    if not record or record.status == 'deleted':
        return False, '文件不存在'

    # Permission check
    if record.scope == 'public' and not _has_public_permission(user, record.directory_id, PERM_DOWNLOAD):
        return False, '没有下载权限'
    elif record.scope == 'private' and record.owner_id != user.id and not user.is_admin():
        return False, '没有权限访问该文件'

    abs_path = _resolve_path(record.file_path)
    if not os.path.exists(abs_path):
        return False, '文件已丢失'

    _audit(user, 'download', f'下载了「{record.original_filename}」')

    return True, send_file(
        abs_path,
        as_attachment=True,
        download_name=record.original_filename,
    )


def list_files(user, scope='public', directory_id=None, page=1, per_page=20,
               keyword=None, file_type=None, uploader=None, sort_by='created_at', sort_order='desc',
               start_date=None, end_date=None):
    """List files with filters."""
    # Build filter conditions once, reuse for both pagination and total_size
    filters = [FileRecord.status == 'active']

    if scope == 'public':
        filters.append(FileRecord.scope == 'public')
    else:
        filters.append(FileRecord.scope == 'private')
        if not user.is_admin():
            filters.append(FileRecord.owner_id == user.id)

    if directory_id:
        filters.append(FileRecord.directory_id == directory_id)

    if keyword:
        like_kw = f'%{keyword.replace("%", "\\%").replace("_", "\\_")}%'
        filters.append(FileRecord.original_filename.like(like_kw))
    if file_type:
        filters.append(FileRecord.file_type == file_type)
    if uploader:
        filters.append(FileRecord.owner.has(User.username == uploader))
    if start_date:
        try:
            from datetime import datetime as dt
            filters.append(FileRecord.created_at >= dt.fromisoformat(start_date))
        except ValueError:
            pass
    if end_date:
        try:
            from datetime import datetime as dt
            filters.append(FileRecord.created_at <= dt.fromisoformat(end_date))
        except ValueError:
            pass

    # Total size across ALL matching files (same filters, no pagination)
    total_size = db.session.query(
        db.func.coalesce(db.func.sum(FileRecord.file_size), 0)
    ).filter(*filters).scalar()

    query = FileRecord.query.filter(*filters)
    # Sorting
    sort_col = FileRecord.file_size if sort_by == 'file_size' else FileRecord.created_at
    order = sort_col.asc() if sort_order == 'asc' else sort_col.desc()
    pagination = query.order_by(order).paginate(
        page=page, per_page=per_page, error_out=False
    )
    return {
        'items': [f.to_dict() for f in pagination.items],
        'total': pagination.total,
        'total_size': total_size,
        'page': page,
        'per_page': per_page,
        'pages': pagination.pages,
    }


def get_filter_options(user, scope='public'):
    """Return distinct file types and uploader names for filter dropdowns."""
    filters = [FileRecord.status == 'active']
    if scope == 'public':
        filters.append(FileRecord.scope == 'public')
    else:
        filters.append(FileRecord.scope == 'private')
        if not user.is_admin():
            filters.append(FileRecord.owner_id == user.id)

    types = (db.session.query(FileRecord.file_type)
             .filter(*filters)
             .distinct()
             .all())
    uploaders = (db.session.query(User.username)
                 .join(FileRecord, FileRecord.owner_id == User.id)
                 .filter(*filters)
                 .distinct()
                 .all())
    return {
        'file_types': sorted([t[0] for t in types if t[0]]),
        'uploaders': sorted([u[0] for u in uploaders if u[0]]),
    }


def delete_file(user, file_id):
    """Soft-delete a file (move to recycle bin)."""
    if not user.is_admin():
        return False, '仅管理员可删除文件'

    record = FileRecord.query.get(file_id)
    if not record or record.status == 'deleted':
        return False, '文件不存在'

    record.status = 'deleted'
    db.session.commit()
    _audit(user, 'delete', f'将「{record.original_filename}」移入回收站')
    return True, '文件已移入回收站'


def list_deleted_files(user, page=1, per_page=20, keyword=None):
    """List files in recycle bin."""
    if not user.is_admin():
        return {'items': [], 'total': 0, 'page': page, 'per_page': per_page, 'pages': 0}
    query = FileRecord.query.filter_by(status='deleted')
    if keyword:
        like_kw = f'%{keyword.replace("%", "\\%").replace("_", "\\_")}%'
        query = query.filter(FileRecord.original_filename.like(like_kw))
    pagination = query.order_by(FileRecord.updated_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False)
    return {
        'items': [f.to_dict() for f in pagination.items],
        'total': pagination.total,
        'page': page,
        'per_page': per_page,
        'pages': pagination.pages,
    }


def restore_file(user, file_id):
    """Restore a file from recycle bin."""
    if not user.is_admin():
        return False, '仅管理员可恢复文件'
    record = FileRecord.query.get(file_id)
    if not record or record.status != 'deleted':
        return False, '文件不在回收站中'
    record.status = 'active'
    db.session.commit()
    _audit(user, 'restore', f'从回收站恢复了「{record.original_filename}」')
    return True, record.to_dict()


def permanent_delete_file(user, file_id):
    """Permanently delete a file from recycle bin (remove physical file)."""
    if not user.is_admin():
        return False, '仅管理员可永久删除文件'
    record = FileRecord.query.get(file_id)
    if not record or record.status != 'deleted':
        return False, '文件不在回收站中'
    # Remove physical file
    if record.file_path:
        abs_path = _resolve_path(record.file_path)
        if os.path.exists(abs_path):
            try:
                os.remove(abs_path)
            except OSError as e:
                return False, f'物理文件删除失败: {e}'

    name = record.original_filename
    _detach_copy_audits(record.id)
    db.session.delete(record)
    db.session.commit()
    _audit(user, 'permanent_delete', f'永久删除了「{name}」')
    return True, '文件已永久删除'


def empty_recycle_bin(user):
    """Permanently delete all files in recycle bin."""
    if not user.is_admin():
        return False, '仅管理员可清空回收站'
    query = FileRecord.query.filter_by(status='deleted')
    records = query.all()
    _detach_copy_audits(*(r.id for r in records))
    deleted = 0
    failed = 0
    for record in records:
        physical_ok = True
        if record.file_path:
            abs_path = _resolve_path(record.file_path)
            if os.path.exists(abs_path):
                try:
                    os.remove(abs_path)
                except OSError:
                    physical_ok = False
                    failed += 1
        if physical_ok:
            db.session.delete(record)
            deleted += 1

    db.session.commit()
    _audit(user, 'empty_recycle_bin',
           f'清空了回收站，永久删除 {deleted} 个文件' + (f'，{failed} 个物理文件删除失败' if failed else ''))
    if failed:
        return True, f'已删除 {deleted} 个记录，{failed} 个物理文件删除失败（记录已保留）'
    return True, f'已永久删除 {deleted} 个文件'


def rename_file(user, file_id, new_name):
    """Rename a file."""
    record = FileRecord.query.get(file_id)
    if not record or record.status == 'deleted':
        return False, '文件不存在'

    if record.scope == 'public' and not user.is_admin():
        return False, '公共目录文件仅管理员可重命名'
    if record.scope == 'private' and record.owner_id != user.id and not user.is_admin():
        return False, '没有权限重命名该文件'

    old_name = record.original_filename
    sanitized = sanitize_filename(new_name) or new_name
    record.original_filename = sanitized
    record.file_type = get_file_extension(new_name)
    db.session.commit()
    _audit(user, 'rename', f'将「{old_name}」重命名为「{sanitized}」')
    return True, record.to_dict()


def _resolve_target_directory(user, record, target_directory_id):
    """Validate and return a writable target directory for move/copy.

    Access control matrix:
    - public scope: target must be public; admin-only write (enforced by caller)
    - private scope: target must be private, owned by the same user as the file;
      non-admin users can only target their own directories.
    Returns (directory, None) on success or (None, error_message) on failure.
    """
    if target_directory_id in (None, ''):
        return None, None

    try:
        target_directory_id = int(target_directory_id)
    except (TypeError, ValueError):
        return None, '目标目录不存在或无权限'

    target_dir = Directory.query.get(target_directory_id)
    if not target_dir or target_dir.scope != record.scope:
        return None, '目标目录不存在或无权限'

    if record.scope == 'private':
        if target_dir.owner_id != record.owner_id:
            return None, '目标目录不存在或无权限'
        if not user.is_admin() and target_dir.owner_id != user.id:
            return None, '目标目录不存在或无权限'

    return target_dir, None


def move_file(user, file_id, target_directory_id):
    """Move a file to another directory."""
    record = FileRecord.query.get(file_id)
    if not record or record.status == 'deleted':
        return False, '文件不存在'

    if record.scope == 'public' and not user.is_admin():
        return False, '仅管理员可移动公共目录文件'
    if record.scope == 'private' and record.owner_id != user.id and not user.is_admin():
        return False, '没有权限移动该文件'

    target_dir, err = _resolve_target_directory(user, record, target_directory_id)
    if err:
        return False, err

    record.directory_id = target_dir.id if target_dir else None
    db.session.commit()
    _audit(user, 'move',
           f'将「{record.original_filename}」移动到「{target_dir.name if target_dir else "根目录"}」')
    return True, record.to_dict()


def copy_file(user, file_id, target_directory_id=None):
    """Copy a file within the system."""
    record = FileRecord.query.get(file_id)
    if not record or record.status == 'deleted':
        return False, '源文件不存在'

    if record.scope == 'public' and not _has_public_permission(user, record.directory_id, PERM_COPY):
        return False, '没有拷贝权限'
    if record.scope == 'private' and record.owner_id != user.id and not user.is_admin():
        return False, '没有权限拷贝该文件'

    target_dir, err = _resolve_target_directory(user, record, target_directory_id)
    if err:
        return False, err

    new_rel_path = _physical_path(record.scope, user, record.original_filename)
    new_abs_path = _resolve_path(new_rel_path)
    shutil.copy2(_resolve_path(record.file_path), new_abs_path)

    new_record = FileRecord(
        filename=os.path.basename(new_rel_path),
        original_filename=record.original_filename,
        file_path=new_rel_path,
        file_size=record.file_size,
        file_type=record.file_type,
        mime_type=record.mime_type,
        directory_id=target_dir.id if target_dir else record.directory_id,
        owner_id=user.id,
        scope=record.scope,
    )
    db.session.add(new_record)
    db.session.flush()

    # Copy audit
    ip = get_client_ip() if request else 'system'
    ua = get_client_ua() if request else ''
    copy_log = CopyAudit(
        user_id=user.id, account=user.account, username=user.username,
        file_id=record.id,
        source_filename=record.original_filename,
        source_path=record.file_path,
        target_path=new_rel_path,
        file_size=record.file_size,
        copy_type='internal', ip=ip, user_agent=ua,
    )
    db.session.add(copy_log)
    db.session.commit()

    _audit(user, 'copy', f'拷贝了「{record.original_filename}」')
    return True, new_record.to_dict()


BATCH_MAX = 100


def _parse_file_ids(file_ids):
    if not file_ids or not isinstance(file_ids, list):
        return None, '请提供文件 ID 列表'
    try:
        ids = list(dict.fromkeys(int(x) for x in file_ids))
    except (TypeError, ValueError):
        return None, '文件 ID 格式无效'
    if not ids:
        return None, '请至少选择一个文件'
    if len(ids) > BATCH_MAX:
        return None, f'单次最多操作 {BATCH_MAX} 个文件'
    return ids, None


def _batch_summary(results):
    success_count = sum(1 for r in results if r['status'] == 'success')
    fail_count = len(results) - success_count
    return {
        'success_count': success_count,
        'fail_count': fail_count,
        'results': results,
    }


def batch_delete_files(user, file_ids):
    """Soft-delete multiple files (move to recycle bin)."""
    ids, err = _parse_file_ids(file_ids)
    if err:
        return False, err
    if not user.is_admin():
        return False, '仅管理员可删除文件'

    results = []
    for file_id in ids:
        record = FileRecord.query.get(file_id)
        if not record or record.status == 'deleted':
            results.append({'file_id': file_id, 'status': 'failure', 'reason': '文件不存在'})
            continue
        record.status = 'deleted'
        results.append({'file_id': file_id, 'filename': record.original_filename, 'status': 'success'})

    db.session.commit()
    summary = _batch_summary(results)
    sc, fc = summary['success_count'], summary['fail_count']
    detail = f'批量移入回收站 {sc} 个文件' + (f'（{fc} 个失败）' if fc else '')
    _audit(user, 'delete', detail)
    return True, summary


def batch_move_files(user, file_ids, target_directory_id):
    """Move multiple files to a directory (None = root)."""
    ids, err = _parse_file_ids(file_ids)
    if err:
        return False, err

    target_dir_name = '根目录'
    results = []
    for file_id in ids:
        record = FileRecord.query.get(file_id)
        if not record or record.status == 'deleted':
            results.append({'file_id': file_id, 'status': 'failure', 'reason': '文件不存在'})
            continue
        if record.scope == 'public' and not user.is_admin():
            results.append({'file_id': file_id, 'filename': record.original_filename,
                            'status': 'failure', 'reason': '仅管理员可移动公共目录文件'})
            continue
        if record.scope == 'private' and record.owner_id != user.id and not user.is_admin():
            results.append({'file_id': file_id, 'filename': record.original_filename,
                            'status': 'failure', 'reason': '没有权限移动该文件'})
            continue
        target_dir, err = _resolve_target_directory(user, record, target_directory_id)
        if err:
            results.append({'file_id': file_id, 'filename': record.original_filename,
                            'status': 'failure', 'reason': err})
            continue
        target_dir_name = target_dir.name if target_dir else '根目录'
        record.directory_id = target_dir.id if target_dir else None
        results.append({'file_id': file_id, 'filename': record.original_filename, 'status': 'success'})

    db.session.commit()
    summary = _batch_summary(results)
    sc, fc = summary['success_count'], summary['fail_count']
    _audit(user, 'move', f'批量移动 {sc} 个文件到「{target_dir_name}」' + (f'（{fc} 个失败）' if fc else ''))
    return True, summary


def batch_copy_files(user, file_ids, target_directory_id=None):
    """Copy multiple files within the system.

    Each file is copied in its own savepoint so a single failure does not
    roll back already-copied files. The final summary audit is written once.
    """
    ids, err = _parse_file_ids(file_ids)
    if err:
        return False, err

    results = []
    for file_id in ids:
        record = FileRecord.query.get(file_id)
        if not record or record.status == 'deleted':
            results.append({'file_id': file_id, 'status': 'failure', 'reason': '源文件不存在'})
            continue
        if record.scope == 'public' and not _has_public_permission(user, record.directory_id, PERM_COPY):
            results.append({'file_id': file_id, 'filename': record.original_filename,
                            'status': 'failure', 'reason': '没有拷贝权限'})
            continue

        target_dir, dir_err = _resolve_target_directory(user, record, target_directory_id)
        if dir_err:
            results.append({'file_id': file_id, 'filename': record.original_filename,
                            'status': 'failure', 'reason': dir_err})
            continue

        # Use a savepoint so a failure only rolls back this single file
        # without affecting previously committed files.
        savepoint = db.session.begin_nested()
        try:
            new_rel_path = _physical_path(record.scope, user, record.original_filename)
            new_abs_path = _resolve_path(new_rel_path)
            shutil.copy2(_resolve_path(record.file_path), new_abs_path)
            new_record = FileRecord(
                filename=os.path.basename(new_rel_path),
                original_filename=record.original_filename,
                file_path=new_rel_path,
                file_size=record.file_size,
                file_type=record.file_type,
                mime_type=record.mime_type,
                directory_id=target_dir.id if target_dir else record.directory_id,
                owner_id=user.id,
                scope=record.scope,
            )
            db.session.add(new_record)
            db.session.flush()

            ip = get_client_ip() if request else 'system'
            ua = get_client_ua() if request else ''
            copy_log = CopyAudit(
                user_id=user.id, account=user.account, username=user.username,
                file_id=record.id,
                source_filename=record.original_filename,
                source_path=record.file_path,
                target_path=new_rel_path,
                file_size=record.file_size,
                copy_type='internal', ip=ip, user_agent=ua,
            )
            db.session.add(copy_log)
            savepoint.commit()  # release savepoint; outer commit at end
            results.append({'file_id': file_id, 'filename': record.original_filename,
                            'status': 'success', 'new_file_id': new_record.id})
        except Exception as e:
            savepoint.rollback()
            # Clean up orphaned physical file if it was created
            if 'new_abs_path' in locals() and os.path.exists(new_abs_path):
                try:
                    os.remove(new_abs_path)
                except OSError:
                    pass
            results.append({'file_id': file_id, 'filename': record.original_filename,
                            'status': 'failure', 'reason': f'复制失败: {e}'})

    db.session.commit()  # single commit for all successful copies
    summary = _batch_summary(results)
    sc, fc = summary['success_count'], summary['fail_count']
    _audit(user, 'copy', f'批量拷贝 {sc} 个文件' + (f'（{fc} 个失败）' if fc else ''))
    return True, summary


def list_directories(user, scope='public', parent_id=None):
    """List directories. Excludes system root directories (/public, /private)."""
    query = Directory.query.filter_by(scope=scope)
    if parent_id:
        query = query.filter_by(parent_id=parent_id)
    else:
        query = query.filter_by(parent_id=None)
        # Exclude system root directories
        query = query.filter(Directory.path != f'/{scope}')

    if scope == 'private' and not user.is_admin():
        query = query.filter_by(owner_id=user.id)

    dirs = query.order_by(Directory.name).all()
    return [d.to_dict() for d in dirs]


def create_directory(user, name, scope='public', parent_id=None):
    """Create a new directory."""
    if scope not in ('public', 'private'):
        return False, 'scope 必须为 public 或 private'
    if scope == 'public' and not user.is_admin():
        return False, '仅管理员可创建公共目录'

    # Sanitize directory name
    name = os.path.basename(name).strip('. ')
    if not name:
        return False, '目录名无效'

    parent_path = f'/{scope}'
    if parent_id:
        parent = Directory.query.get(parent_id)
        if not parent:
            return False, '父目录不存在'
        if parent.scope != scope:
            return False, '父目录的范围与新目录不匹配'
        parent_path = parent.path

    path = f"{parent_path.rstrip('/')}/{name}"
    if Directory.query.filter_by(path=path).first():
        return False, f'目录 "{name}" 已存在'

    d = Directory(
        name=name, parent_id=parent_id, scope=scope,
        owner_id=user.id if scope == 'private' else None,
        path=path, created_by=user.id,
    )
    db.session.add(d)
    db.session.commit()
    parent_name = '根目录'
    if parent_id:
        p = Directory.query.get(parent_id)
        if p:
            parent_name = p.name
    _audit(user, 'create_directory', f'在「{parent_name}」下创建了目录「{name}」')
    return True, d.to_dict()


def delete_directory(user, dir_id):
    """Delete an empty directory."""
    d = Directory.query.get(dir_id)
    if not d:
        return False, '目录不存在'
    if d.scope == 'public' and not user.is_admin():
        return False, '仅管理员可删除公共目录'
    if d.children.count() > 0:
        return False, '目录不为空，无法删除'
    if d.files.filter_by(status='active').count() > 0:
        return False, '目录下有文件，无法删除'

    db.session.delete(d)
    db.session.commit()
    _audit(user, 'delete_directory', f'删除了目录「{d.name}」')
    return True, '目录已删除'


def rename_directory(user, dir_id, new_name):
    """Rename a directory."""
    d = Directory.query.get(dir_id)
    if not d:
        return False, '目录不存在'
    if d.scope == 'public' and not user.is_admin():
        return False, '仅管理员可重命名公共目录'
    if d.scope == 'private' and d.owner_id != user.id and not user.is_admin():
        return False, '没有权限重命名该目录'

    # Sanitize name
    new_name = os.path.basename(new_name).strip('. ')
    if not new_name:
        return False, '目录名无效'
    if new_name == d.name:
        return True, d.to_dict()

    # Update path: replace old name with new name in the full path
    old_path = d.path
    parent_prefix = old_path.rsplit('/', 1)[0] if '/' in old_path[1:] else ''
    new_path = f'{parent_prefix}/{new_name}'
    if Directory.query.filter_by(path=new_path).first():
        return False, f'目录 "{new_name}" 已存在'

    old_name = d.name
    d.name = new_name
    d.path = new_path

    # Update children paths recursively
    _update_children_paths(d, old_path, new_path)

    db.session.commit()
    _audit(user, 'rename_directory', f'将目录「{old_name}」重命名为「{new_name}」')
    return True, d.to_dict()


def _update_children_paths(directory, old_prefix, new_prefix):
    """Recursively update path prefixes for all descendant directories."""
    children = Directory.query.filter(
        Directory.path.like(f'{old_prefix}/%')
    ).all()
    for child in children:
        child.path = new_prefix + child.path[len(old_prefix):]


def get_file_preview(user, file_id):
    """Get file for preview (inline display).

    Conversion strategy:
    - Word (.docx/.doc) → simplified HTML (headings + paragraphs only, no images/tables)
    - Excel (.xlsx/.xls) → HTML table (values only, no formulas or formatting)
    - Everything else → raw file download with original MIME type

    All user-generated content is escaped with markupsafe.escape() to prevent XSS.
    """
    record = FileRecord.query.get(file_id)
    if not record or record.status == 'deleted':
        return False, '文件不存在'

    if record.scope == 'public' and not _has_public_permission(user, record.directory_id, PERM_DOWNLOAD):
        return False, '没有预览权限'
    if record.scope == 'private' and record.owner_id != user.id and not user.is_admin():
        return False, '没有权限访问该文件'

    abs_path = _resolve_path(record.file_path)
    if not os.path.exists(abs_path):
        return False, '文件已丢失'

    _audit(user, 'preview', f'预览了「{record.original_filename}」')

    ext = os.path.splitext(record.original_filename)[1].lower()

    # Word → simplified HTML (headings + paragraphs, no images/tables).
    # python-docx only supports .docx; .doc is not supported.
    if ext == '.docx':
        try:
            from docx import Document
            doc = Document(abs_path)
            html = '<div style="max-width:900px;margin:0 auto;padding:24px;font-family:sans-serif;color:#1E293B;">'
            for p in doc.paragraphs:
                if p.style.name.startswith('Heading'):
                    level = int(p.style.name[-1]) if p.style.name[-1].isdigit() else 1
                    sizes = {1:24, 2:20, 3:16, 4:14}
                    html += f'<h{level} style="font-size:{sizes.get(level,16)}px;margin:0.5em 0;font-weight:600;">{escape(p.text)}</h{level}>'
                elif p.text.strip():
                    html += f'<p style="margin:0.4em 0;line-height:1.7;">{escape(p.text)}</p>'
            html += '</div>'
            from flask import make_response
            resp = make_response(html)
            resp.headers['Content-Type'] = 'text/html; charset=utf-8'
            return True, resp
        except ImportError:
            return False, 'Word 预览功能依赖缺失，请联系管理员'
        except Exception as e:
            logger.warning(f"Word preview failed for {record.file_path}: {e}")
            return False, 'Word 文件预览失败，文件可能已损坏'

    if ext == '.doc':
        return False, '.doc 格式暂不支持预览，请转换为 .docx 格式或下载后查看'

    # Excel → HTML table (values only, no formulas or formatting).
    # openpyxl only supports .xlsx; .xls is not supported.
    if ext == '.xlsx':
        try:
            from openpyxl import load_workbook
            wb = load_workbook(abs_path, data_only=True)
            sheet = wb.active
            html = '<div style="max-width:100%;overflow:auto;padding:16px;">'
            html += '<table style="border-collapse:collapse;font-family:sans-serif;font-size:13px;">'
            for row in sheet.iter_rows(values_only=True):
                html += '<tr>'
                for cell in row:
                    val = str(cell) if cell is not None else ''
                    html += f'<td style="border:1px solid #D5D8DC;padding:6px 10px;min-width:80px;color:#1E293B;">{escape(val)}</td>'
                html += '</tr>'
            html += '</table></div>'
            from flask import make_response
            resp = make_response(html)
            resp.headers['Content-Type'] = 'text/html; charset=utf-8'
            return True, resp
        except ImportError:
            return False, 'Excel 预览功能依赖缺失，请联系管理员'
        except Exception as e:
            logger.warning(f"Excel preview failed for {record.file_path}: {e}")
            return False, 'Excel 文件预览失败，文件可能已损坏'

    if ext == '.xls':
        return False, '.xls 格式暂不支持预览，请转换为 .xlsx 格式或下载后查看'

    return True, send_file(
        abs_path,
        mimetype=record.mime_type or 'application/octet-stream',
    )
